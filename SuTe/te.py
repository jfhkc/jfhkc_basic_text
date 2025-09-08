from openpyxl import *
from openpyxl.drawing.image import Image
from openpyxl.cell import MergedCell
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D

from io import BytesIO
import requests
from tqdm import tqdm
import traceback


class ExcelOp(object):
    def __init__(self, file, sheet_name="Sheet1"):
        self.file = file
        self.wb = load_workbook(self.file)
        self.ws = self.wb[sheet_name]
        # self.max_rows, self.max_cols = self.ws.max_row, self.ws.max_column

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell = self._parser_merged_cell(row, column)
        cell_value = cell.value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.get_cell_value(row=i, column=column)
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.get_cell_value(row=row, column=i)
            row_data.append(cell_value)
        return row_data

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        cell = self._parser_merged_cell(row, colunm)
        try:
            cell.value = cellvalue
        except:
            cell.value = "ERROR:writefail"

    # 在某个单元格上添加一张图
    def set_image(self, fp: BytesIO, row: int, column: int, img_pixel_height=None, img_pixel_width=None):
        cell = self._parser_merged_cell(row, column)
        image_data = Image(fp)
        w = image_data.width if img_pixel_width is None else img_pixel_width
        h = image_data.height if img_pixel_height is None else img_pixel_height
        size = XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h))
        # https://stackoverflow.com/questions/55309671/more-precise-image-placement-possible-with-openpyxl-pixel-coordinates-instead
        # AnchorMarker 它的row和col 又从0开始数了 好烦
        marker = AnchorMarker(col=cell.column-1, row=cell.row-1)
        image_data.anchor = OneCellAnchor(_from=marker, ext=size)
        # image_data.anchor = "A1"
        self.ws.add_image(image_data)

    def _parser_merged_cell(self, row, col):
        """
        检查是否为合并单元格并获取对应行列单元格的值。
        如果是合并单元格，则取合并区域左上角单元格的值作为当前单元格的值,否则直接返回该单元格的值
        :param sheet: 当前工作表对象
        :param row: 需要获取的单元格所在行
        :param col: 需要获取的单元格所在列
        :return: 
        """
        cell = self.ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
            for merged_range in self.ws.merged_cells.ranges:  # 循环查找该单元格所属的合并区域
                if cell.coordinate in merged_range:
                    # 获取合并区域左上角的单元格作为该单元格的值返回
                    cell = self.ws.cell(
                        row=merged_range.min_row, column=merged_range.min_col)
                    break
        return cell


if __name__ == "__main__":
    excel = ExcelOp(file="D:/JT_t.xlsx")
    excel.ws.column_dimensions['E'].width = 50  # 设置单元格宽
    urls = excel.get_col_value(5)  # excel行和列都是从1开始数  7对应G
    try:
        for row, url in tqdm(enumerate(urls[1:])):
            row = row+2  # python从0开始数  excel从1开始  且excel第一行是title 所以加2
            # if url.endswith('_30x30.jpg'):
            #     url = url[0:-10]
            if url is not None and url.endswith('_30x30.jpg'):
                url = url[:-10]  # 注意：Python 允许负索引，所以 url[:-10] 和 url[0:-10] 效果相同
            excel.ws.row_dimensions[row].height = 50  # 设置单元格高
            excel.set_cell_value(row, 5, '')  # 删除原有的url
            r = requests.get(url=url, headers={"user-agent": "Mozilla/5.0"})
            if r.status_code != 200:
                print(r.status_code, " in line ", row)
            # 文件操作open()的返回值就是BytesIO 或 StringIO
            excel.set_image(BytesIO(r.content), row, 5, 305.5, 405.8)
    except:
        print(traceback.format_exc())
        excel.wb.save("2_witherror.xlsx")
    excel.wb.save("2.xlsx")

